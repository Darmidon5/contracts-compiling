from openpyxl import load_workbook
from docxtpl import DocxTemplate
from json import loads
from counterparty import Counterparty


def get_paths():
    with open('json_data', 'r', encoding='utf-8') as file:
        paths = loads(file.readline().rstrip())
        xlpath = paths['xl']
        contractpath = paths['conctract']
        templatepath = paths['template']
    return xlpath, contractpath, templatepath


def get_xl_data(xlpath):
    workbook = load_workbook(rf'{xlpath}.xlsx')
    sheet_1 = workbook['Лист1']
    return sheet_1


def content_gathering(sheet):
    customers_data = []
    columns = ['id', 'name', 'contract_date', 'representative_post', 'representative_name', 'representative_document',
               'document_number', 'document_date', 'rental_period', 'rental_cost', 'VAT', 'security_payment',
               'authorized_to_transfer_property', 'authorized_to_return_property', 'purposes_of_use',
               'operating_address', 'address', 'post_address', 'phone_number', 'fax', 'email', 'OGRN', 'INN', 'KPP',
               'payment_account', 'bank_name', 'bank_name', 'correspondent_account', 'BIK']
    for i in range(3, sheet.max_row + 1):
        company_data = dict()
        for idx in range(1, len(columns)+1):
            value = sheet.cell(i, idx).value
            if value:
                company_data[columns[idx-1]] = sheet.cell(i, idx).value

        customers_data.append(Counterparty(company_data))

    return customers_data


def template_filling(templatepath, contractpath, customers_data):
    template = DocxTemplate(rf'{templatepath}.docx')
    for customer in customers_data:
        context = customer.__dict__
        template.render(context)
        # C:\Users\niks_\Desktop\тестим договоры
        template.save(rf'{contractpath}\{customer["name"]} договор.docx')


def compile_contracts():
    xlpath, contractpath, templatepath = get_paths()
    sheet = get_xl_data(xlpath)
    customers_data = content_gathering(sheet)
    template_filling(templatepath, contractpath, customers_data)
